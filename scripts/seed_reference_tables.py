#!/usr/bin/env python3
"""Seed the read-only reference tables from the PIER lead-lake workbook.

Loads:
  * "Pier Pipeline"   sheet -> public.pier_pipeline
  * "EUREFAS Members" sheet -> public.eurefas_members

Both sheets use header row 4 and data from row 5 onward.

These two tables are read-only reference tables: RLS is enabled but there is no
INSERT policy, so writes must go through the Supabase *service role* key (which
bypasses RLS). Set these environment variables before running:

    export SUPABASE_URL="https://<project-ref>.supabase.co"
    export SUPABASE_SERVICE_ROLE_KEY="<service-role-key>"

Then:

    python scripts/seed_reference_tables.py            # load into Supabase
    python scripts/seed_reference_tables.py --dry-run  # parse + print counts only

The load is idempotent: rows are upserted on their natural unique key
(pipeline_id / member_id), so re-running updates in place instead of duplicating.

Dependencies: openpyxl, supabase  (pip install openpyxl supabase)
"""
from __future__ import annotations

import argparse
import os
import sys

from openpyxl import load_workbook

# Default workbook location (override with --workbook).
DEFAULT_WORKBOOK = os.path.expanduser(
    "~/Documents/260427_PIER_lead_lake_v09_OM_C2.xlsx"
)

# Data starts on row 5 (row 4 is the header) on both sheets.
DATA_START_ROW = 5


def _clean(value):
    """Trim strings; treat blank strings as NULL. Non-strings pass through."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def read_pier_pipeline(ws) -> list[dict]:
    """Pier Pipeline sheet: Pipeline ID, Company Name, Stage, Product Line, Source Folder."""
    rows: list[dict] = []
    for pipeline_id, company_name, stage, product_line, source_folder in (
        r[:5] for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True)
    ):
        pipeline_id = _clean(pipeline_id)
        company_name = _clean(company_name)
        # Skip fully blank / spacer rows.
        if not pipeline_id and not company_name:
            continue
        rows.append(
            {
                "pipeline_id": pipeline_id,
                "company_name": company_name,
                "stage": _clean(stage),
                "product_line": _clean(product_line),
                "source_folder": _clean(source_folder),
            }
        )
    return rows


def read_eurefas_members(ws) -> list[dict]:
    """EUREFAS Members sheet: Member ID, Company Name, Membership, Country, Notes."""
    rows: list[dict] = []
    for member_id, company_name, membership, country, notes in (
        r[:5] for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True)
    ):
        member_id = _clean(member_id)
        company_name = _clean(company_name)
        if not member_id and not company_name:
            continue
        rows.append(
            {
                "member_id": member_id,
                "company_name": company_name,
                "membership": _clean(membership),
                "country": _clean(country),
                "notes": _clean(notes),
            }
        )
    return rows


def load(table: str, rows: list[dict], conflict_key: str, client) -> int:
    """Upsert rows into `table` on `conflict_key`; return the number of rows sent."""
    if not rows:
        return 0
    client.table(table).upsert(rows, on_conflict=conflict_key).execute()
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="path to the .xlsx workbook")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="parse the workbook and print counts without touching the database",
    )
    args = parser.parse_args()

    if not os.path.exists(args.workbook):
        print(f"ERROR: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    pier_rows = read_pier_pipeline(wb["Pier Pipeline"])
    eurefas_rows = read_eurefas_members(wb["EUREFAS Members"])

    print(f"Parsed 'Pier Pipeline':   {len(pier_rows)} rows")
    print(f"Parsed 'EUREFAS Members': {len(eurefas_rows)} rows")

    if args.dry_run:
        print("--dry-run: no rows written.")
        return 0

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print(
            "ERROR: set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY to load "
            "(reference tables have no INSERT policy; the service role is required). "
            "Use --dry-run to parse without a database.",
            file=sys.stderr,
        )
        return 2

    from supabase import create_client  # imported here so --dry-run needs no supabase-py

    client = create_client(url, key)

    pier_n = load("pier_pipeline", pier_rows, "pipeline_id", client)
    eurefas_n = load("eurefas_members", eurefas_rows, "member_id", client)

    print("\nInserted / upserted per table:")
    print(f"  public.pier_pipeline:   {pier_n}")
    print(f"  public.eurefas_members: {eurefas_n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
