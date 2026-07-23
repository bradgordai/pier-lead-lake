#!/usr/bin/env python3
"""Load Companies / Contacts / Outreach Log from the PIER lead-lake workbook.

Sheets use header row 4, data from row 5 (same as the reference-table seeder).

    python scripts/load_full_data.py --phase a --emit-sql out.sql   # first 30 companies
    python scripts/load_full_data.py --phase b --emit-sql out.sql   # everything
    python scripts/load_full_data.py --phase a --dry-run            # parse + report only

Phase A loads the first 30 companies so the Lovable UI can be validated before
committing to the full load. Phase B loads all companies, contacts and outreach.

The emitted SQL resolves foreign keys with subqueries rather than hardcoded UUIDs:
  team_id     -> (SELECT id FROM public.teams WHERE slug = 'pier')
  company_id  -> (SELECT id FROM public.companies WHERE company_id = '<Cnnn>')
  contact_id  -> (SELECT id FROM public.contacts  WHERE contact_id = '<Pnnn>')

Columns deliberately NOT written (database triggers own them):
  companies.tracking                 - tg_resolve_tracking cascade
  companies.contacts_count           - tg_recompute_contacts_count (would double count)
  companies.root_domain              - tg_normalise_company_fields
  contacts.email_normalised          - tg_normalise_contact_fields
  contacts.cooldown_status_derived   - tg_update_cooldown_status

Dependencies: openpyxl (plus supabase only if loading directly rather than emitting SQL).
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import sys
from collections import Counter

from openpyxl import load_workbook

DEFAULT_WORKBOOK = os.path.expanduser("~/Documents/260427_PIER_lead_lake_v09_OM_C2.xlsx")
DATA_START_ROW = 5

# Phase A is an explicit, priority-weighted sample rather than the first N rows.
# Sheet order put 28 of the first 30 in OoS with sparse fields, which barely
# exercised the UI. This set is top-of-band across every priority so the Companies
# list is validated against richly-populated rows (revenue, employees, monthly
# visits, category arrays, insurance state, notes).
# Distribution: 10 P0, 8 P1, 6 P2, 4 P3, 2 OoS.
PHASE_A_COMPANY_IDS = [
    # P0
    "C218", "C263", "C223", "C229", "C119", "C045", "C185", "C207", "C001", "C148",
    # P1
    "C241", "C093", "C246", "C109", "C242", "C061", "C216", "C295",
    # P2
    "C238", "C285", "C290", "C329", "C330", "C257",
    # P3
    "C208", "C132", "C140", "C255",
    # OoS
    "C350", "C302",
]

warnings: Counter = Counter()


def warn(field: str, value) -> None:
    warnings[f"{field}: {value!r}"] += 1


# --------------------------------------------------------------------------
# Enum members (must match migration 002 exactly)
# --------------------------------------------------------------------------
E_PRIORITY = {"P0", "P1", "P2", "P3", "OoS", "Competitor"}
E_RESEARCH = {"Untouched", "Light triage", "Deep research done", "Outdated"}
E_OPPORTUNITY = {"To Review", "Prospect", "Contacted", "Active Lead", "Partner", "Out of Scope"}
E_CATEGORY = {
    "Pure Online Phone Retailer", "Refurbished Specialist", "Electronics",
    "Multi-Category Retailer", "Operator", "Manufacturer", "Marketplace",
    "Comparison Site", "Industry Media", "Influencer", "Other",
}
E_INS_STRUCTURE = {
    "Optional Add-On", "Bundled", "Upsold", "Embedded in T&Cs",
    "Redirect to Third-Party", "Other",
}
E_INDUSTRY = {
    "Mobile/Gadget Retail", "Refurb / Recommerce", "Telco", "Manufacturer",
    "Software", "Telco Infrastructure", "Industry Media", "Influencer", "Other",
}
E_PRODUCT_LINE = {"Pier Protect", "Ticketplan", "TIGA", "Multiple", "Unknown"}
E_ACCOUNT_OWNER = {"Oliver Müller", "Phil", "Mark"}
E_SENIORITY = {"C-suite", "Senior", "Director", "Manager", "Other"}
E_FUNCTION = {
    "Alliances / BD", "Marketing", "Product", "Engineering", "Sales", "Finance",
    "Operations", "Legal", "HR", "Executive", "Other",
    "Strategy",  # added in migration 018
}
E_CONNECTION_LEVEL = {"1st degree", "2nd degree", "3rd degree", "Not connected"}
E_FORMALITY = {"Formal", "Informal"}
E_LANGUAGE = {"EN", "DE", "FR", "ES", "IT", "NL", "Other"}
E_CONN_STATUS = {
    "Not connected", "Request sent", "Accepted", "Already connected", "Ignored", "Withdrawn",
}
E_OUTREACH_STATUS = {
    "Not started", "Ready", "Active", "Contacted", "In conversation", "Cooldown",
    "Needs review", "Do not contact", "Left company",
    "Not relevant",  # added in migration 018
}
E_CHANNEL = {
    "LinkedIn DM", "LinkedIn CR", "LinkedIn inMail", "Email", "Phone", "In-person", "Other",
}
E_TOUCH_TYPE = {
    "Initial message", "Connection request", "Chase", "Reply", "Event follow-up",
    "Introduction", "Meeting confirmation", "Other",
}
E_SEND_STATUS = {"Draft", "Ready", "Scheduled", "Sent", "Cancelled"}
E_OUTCOME = {
    "Awaiting reply", "Replied / Accepted", "No reply", "Rejected / Bounced", "Withdrawn",
}
E_REPLY_CLASS = {
    "Positive interest", "Neutral", "Objection", "Not interested", "Out of office",
    "Wrong person", "Do not contact", "Booked meeting", "Uncategorised",
}

# --------------------------------------------------------------------------
# Aliases: workbook value -> enum member.
#
# Anything not covered here and not an exact enum match is dropped with a warning
# (or falls back to the column default for NOT NULL columns).
#
# Two workbook values that were previously aliased are now first-class enum
# members (migration 018), so they map to themselves and are no longer coerced:
#   'Strategy'      contacts.function        was -> 'Executive'      (78 rows)
#   'Not relevant'  contacts.outreach_status was -> 'Do not contact' (8 rows)
# 'Do not contact' is a GDPR/consent suppression state, so conflating it with a
# commercial "not relevant" judgement was both lossy and unsafe.
#
# Remaining aliases below are still genuine coercions and are worth reviewing:
#   'Commercial / Sales' -> 'Sales'            (45 contacts)
#   'VP / Director'      -> 'Director'         (27 contacts)
#   'Not sent'           -> 'Not connected'    (54 contacts)
#   'To contact'         -> 'Not started'      (55 contacts)
#   'Accepted'/'Replied' -> 'Replied / Accepted' (38 touches)
#   reply-type variants  -> 'Reply'            (~20 touches)
# --------------------------------------------------------------------------
A_INDUSTRY = {
    "Software Provider": "Software",
    "Telco / Manufacturer": "Telco",
    "Software / Telco Infrastructure": "Telco Infrastructure",
    "Industry Media / Influencer": "Industry Media",
    "B2B IT Leasing / Asset Management": "Other",
}
A_CATEGORY = {  # applied per split token
    "Refurb": None, "Recommerce": None,
    "Software Provider": None, "Wholesaler": None, "Distributor": None,
}
A_SENIORITY = {"VP / Director": "Director", "VP": "Director", "Mid": "Other", "IC": "Other"}
A_FUNCTION = {
    "Strategy": "Strategy",  # identity since migration 018 (was -> 'Executive')
    "Innovation / Strategy": "Strategy", "Leadership": "Executive",
    "C-suite": "Executive", "Commercial / Sales": "Sales", "Insurance / Risk": "Other",
    "Procurement": "Operations",
}
A_CONNECTION_LEVEL = {"Unknown": None}
A_FORMALITY = {"Sie / formal": "Formal"}
A_CONN_STATUS = {"Not sent": "Not connected"}
A_OUTREACH_STATUS = {
    "To contact": "Not started",
    "Not relevant": "Not relevant",  # identity since migration 018 (was -> 'Do not contact')
}
A_CHANNEL = {"Sales Nav InMail": "LinkedIn inMail"}
A_TOUCH_TYPE = {
    "Reply received": "Reply", "Reply to no": "Reply", "Reply to referral": "Reply",
    "Reply ack": "Reply", "Reply to question": "Reply", "Inbound reply": "Reply",
    "Soft-Welcome": "Introduction", "Active-client touch": "Other",
    "Initial email": "Initial message", "CR-accept": "Connection request",
}
A_SEND_STATUS = {"Received": "Sent"}
A_OUTCOME = {
    "Accepted": "Replied / Accepted", "Replied": "Replied / Accepted",
    "No response": "No reply",
}


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def strip_paren(s):
    """'Prospect (relationship/PR play)' -> 'Prospect'; also drops trailing notes."""
    if not isinstance(s, str):
        return s
    s = s.split("\n")[0].strip()
    if "(" in s:
        s = s.split("(")[0].strip()
    return s or None


def enum_val(raw, members, aliases=None, field="", fallback=None):
    """Map a workbook value onto an enum member, else warn and return fallback."""
    v = clean(raw)
    if v is None:
        return fallback
    if isinstance(v, str):
        if v in members:
            return v
        if aliases and v in aliases:
            return aliases[v]
        base = strip_paren(v)
        if base in members:
            return base
        if aliases and base in aliases:
            return aliases[base]
    warn(field, v)
    return fallback


def enum_array(raw, members, aliases, field):
    """Multi-select: split on ';' or '/', map each token, drop unknowns."""
    v = clean(raw)
    if v is None:
        return []
    v = strip_paren(v) or ""
    tokens = [t.strip() for part in v.split(";") for t in part.split("/")]
    out = []
    for t in tokens:
        if not t:
            continue
        if t in members:
            if t not in out:
                out.append(t)
        elif aliases and t in aliases and aliases[t]:
            if aliases[t] not in out:
                out.append(aliases[t])
        elif aliases and t in aliases:
            warn(field, t)  # explicitly known-unmappable
        else:
            warn(field, t)
    return out


EXCEL_EPOCH = dt.date(1899, 12, 30)  # Excel's day 0 (1900 leap-year bug accounted for)


def as_date(v, field=""):
    """Accept datetime/date, ISO 'YYYY-MM-DD' strings, and Excel serial numbers."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, (int, float)):  # Excel serial date
        try:
            return (EXCEL_EPOCH + dt.timedelta(days=int(v))).isoformat()
        except (OverflowError, ValueError):
            warn(field + " (bad serial date)", v)
            return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return dt.datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    warn(field + " (not a date)", v)
    return None


def as_num(v, field=""):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    warn(field + " (not numeric)", v)
    return None


def as_int(v, field=""):
    n = as_num(v, field)
    if n is None:
        return None
    try:
        return int(n)
    except (TypeError, ValueError):
        warn(field + " (not int)", v)
        return None


def as_bool(v):
    v = clean(v)
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in {"yes", "true", "y", "1"}


# --------------------------------------------------------------------------
# Sheet readers
# --------------------------------------------------------------------------
def read_companies(ws):
    out = []
    for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        r = list(r) + [None] * (39 - len(r))
        if not clean(r[0]):
            continue
        out.append({
            "company_id": clean(r[0]),                                  # A
            "company_name": clean(r[1]),                                # B
            # C Tracking -> trigger owns it
            "priority": enum_val(r[3], E_PRIORITY, None, "companies.priority"),
            "research_stage": enum_val(r[4], E_RESEARCH, None, "companies.research_stage", "Untouched"),
            # F Contacts -> trigger owns it
            "website_url": clean(r[6]),
            "country": clean(r[7]),
            "category": enum_array(r[8], E_CATEGORY, A_CATEGORY, "companies.category"),
            "refurbished_offered": clean(r[9]),
            "sim_free_devices": clean(r[10]),
            "parent_group": clean(r[11]),
            "headquarter_location": clean(r[12]),
            "countries_selling_in": clean(r[13]),
            "estimated_revenue_gbp": as_num(r[14], "companies.estimated_revenue_gbp"),
            "employees": as_int(r[15], "companies.employees"),
            "monthly_visits": as_int(r[16], "companies.monthly_visits"),
            "creditsafe_rating": clean(r[17]),
            "insurance_offered": clean(r[18]),
            "insurance_provider": clean(r[19]),
            "insurance_product_types": [t.strip() for t in str(clean(r[20]) or "").split(";") if t.strip()],
            "insurance_structure_type": enum_val(r[21], E_INS_STRUCTURE, None, "companies.insurance_structure_type"),
            "insurance_monthly_price": as_num(r[22], "companies.insurance_monthly_price"),
            "insurance_annual_price": as_num(r[23], "companies.insurance_annual_price"),
            "distribution_model": clean(r[24]),
            "coverage_summary": clean(r[25]),
            "customer_journey": clean(r[26]),
            "policy_url": clean(r[27]),
            "opportunity_status": enum_val(r[28], E_OPPORTUNITY, None, "companies.opportunity_status", "To Review"),
            "usp_notes": clean(r[29]),
            "additional_notes": clean(r[30]),
            "industry": enum_val(r[31], E_INDUSTRY, A_INDUSTRY, "companies.industry"),
            "product_line": enum_val(r[32], E_PRODUCT_LINE, None, "companies.product_line", "Unknown"),
            "account_owner": enum_val(r[33], E_ACCOUNT_OWNER, None, "companies.account_owner"),
            "account_source": clean(r[34]),
            "last_refreshed": as_date(r[35], "companies.last_refreshed"),
            "source_urls": clean(r[36]),
            "annual_devices_sold": clean(r[37]) and str(clean(r[37])),
            "date_added": as_date(r[38], "companies.date_added"),
            "legacy_source": "excel_v09",
        })
    return out


def read_contacts(ws):
    out = []
    for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        r = list(r) + [None] * (31 - len(r))
        if not clean(r[0]):
            continue
        out.append({
            "contact_id": clean(r[0]),          # A
            "company_ref": clean(r[1]),         # B (Cnnn)
            # C Company Name, D Lead Priority are VLOOKUPs, resolved by join
            "first_name": clean(r[4]),
            "last_name": clean(r[5]),
            "job_title": clean(r[6]),
            "seniority": enum_val(r[7], E_SENIORITY, A_SENIORITY, "contacts.seniority"),
            "function": enum_val(r[8], E_FUNCTION, A_FUNCTION, "contacts.function"),
            "location": clean(r[9]),
            "linkedin_url": clean(r[10]),
            "linkedin_sales_nav_url": clean(r[11]),
            "email": clean(r[12]),
            "phone": clean(r[13]) and str(clean(r[13])),
            "connection_level": enum_val(r[14], E_CONNECTION_LEVEL, A_CONNECTION_LEVEL, "contacts.connection_level"),
            "formality": enum_val(r[15], E_FORMALITY, A_FORMALITY, "contacts.formality"),
            "language_code": enum_val(r[16], E_LANGUAGE, None, "contacts.language_code"),
            "source_list": clean(r[17]),
            "connection_status": enum_val(r[18], E_CONN_STATUS, A_CONN_STATUS, "contacts.connection_status", "Not connected"),
            "outreach_status": enum_val(r[19], E_OUTREACH_STATUS, A_OUTREACH_STATUS, "contacts.outreach_status", "Not started"),
            "last_contacted": as_date(r[20], "contacts.last_contacted"),
            "next_action": clean(r[21]),
            "next_action_date": as_date(r[22], "contacts.next_action_date"),
            "background_notes": clean(r[23]),
            "country": clean(r[24]),
            "city": clean(r[25]),
            "cooldown_until": as_date(r[26], "contacts.cooldown_until"),
            "do_not_contact": as_bool(r[27]),
            # AC Cooldown Status -> trigger owns it
            "date_added": as_date(r[29], "contacts.date_added"),
            "sn_lists": [t.strip() for t in str(clean(r[30]) or "").split(";") if t.strip()],
            "legacy_source": "excel_v09",
        })
    return out


def read_outreach(ws):
    out = []
    for r in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        r = list(r) + [None] * (18 - len(r))
        if not clean(r[0]):
            continue
        out.append({
            "touch_id": clean(r[0]),        # A
            "contact_ref": clean(r[1]),     # B (Pnnn)
            # C Contact Name, D Company Name, E/F/G LinkedIn+Email are VLOOKUPs
            "touch_date": as_date(r[7], "outreach.touch_date"),
            "channel": enum_val(r[8], E_CHANNEL, A_CHANNEL, "outreach.channel", "Other"),
            "sent_by": clean(r[9]),
            "touch_type": enum_val(r[10], E_TOUCH_TYPE, A_TOUCH_TYPE, "outreach.touch_type", "Other"),
            "message_body": clean(r[11]),
            "send_status": enum_val(r[12], E_SEND_STATUS, A_SEND_STATUS, "outreach.send_status", "Draft"),
            "outcome": enum_val(r[13], E_OUTCOME, A_OUTCOME, "outreach.outcome"),
            "next_action": clean(r[14]),
            "next_action_date": as_date(r[15], "outreach.next_action_date"),
            # Q Reply Classification in the workbook is free-text narrative, not an
            # enum member. Preserved verbatim in reply_content; classification left
            # NULL rather than forced into a wrong bucket.
            "reply_classification": None,
            "reply_content": clean(r[16]),
            "subject_line": clean(r[17]),
            "migrated_legacy": True,
            "pre_lint_pass": None,
            "legacy_source": "excel_v09",
        })
    return out


# --------------------------------------------------------------------------
# SQL emission
# --------------------------------------------------------------------------
def q(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, list):
        if not v:
            return "'{}'"
        inner = ",".join('"' + str(x).replace("\\", "\\\\").replace('"', '\\"') + '"' for x in v)
        return "'{" + inner.replace("'", "''") + "}'"
    return "'" + str(v).replace("'", "''") + "'"


TEAM = "(SELECT id FROM public.teams WHERE slug = 'pier')"

# NOT NULL columns carrying a DEFAULT: emitting explicit NULL would override the
# default and violate the constraint, so emit the default expression instead.
NOT_NULL_DEFAULTS = {"date_added": "CURRENT_DATE"}


def q_col(col, val):
    if val is None and col in NOT_NULL_DEFAULTS:
        return NOT_NULL_DEFAULTS[col]
    return q(val)


def sql_companies(rows):
    cols = [c for c in rows[0].keys()]
    lines = []
    for r in rows:
        vals = ",".join(q_col(c, r[c]) for c in cols)
        lines.append(f"({TEAM},{vals})")
    collist = ",".join(cols)
    return (
        f"INSERT INTO public.companies (team_id,{collist}) VALUES\n"
        + ",\n".join(lines)
        + "\nON CONFLICT (company_id) DO NOTHING;"
    )


def sql_contacts(rows):
    cols = [c for c in rows[0].keys()]
    lines = []
    for r in rows:
        vals = ",".join(q_col(c, r[c]) for c in cols)
        ref = q(r["company_ref"])
        fk = f"(SELECT id FROM public.companies WHERE company_id = {ref})"
        lines.append(f"({TEAM},{fk},{vals})")
    collist = ",".join(cols)
    return (
        f"INSERT INTO public.contacts (team_id,company_id,{collist}) VALUES\n"
        + ",\n".join(lines)
        + "\nON CONFLICT (contact_id) DO NOTHING;"
    )


def sql_outreach(rows):
    cols = [c for c in rows[0].keys()]
    lines = []
    for r in rows:
        vals = ",".join(q_col(c, r[c]) for c in cols)
        ref = q(r["contact_ref"])
        cfk = f"(SELECT id FROM public.contacts WHERE contact_id = {ref})"
        cofk = (
            "(SELECT c.company_id FROM public.contacts c WHERE c.contact_id = " + ref + ")"
        )
        lines.append(f"({TEAM},{cfk},{cofk},{vals})")
    collist = ",".join(cols)
    return (
        f"INSERT INTO public.outreach_log (team_id,contact_id,company_id,{collist}) VALUES\n"
        + ",\n".join(lines)
        + "\nON CONFLICT (touch_id) DO NOTHING;"
    )


def _strip_none(d):
    """Drop keys whose value is None so DB defaults (e.g. date_added=CURRENT_DATE) apply."""
    return {k: v for k, v in d.items() if v is not None}


def _chunk(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def direct_load(companies, contacts, outreach, url, key):
    """Load via supabase-py (service role bypasses RLS). Resolves team_id and all
    FK UUIDs client-side, FK-validates contacts/outreach, writes reject CSVs, and
    reports counts + timing. Idempotent via upsert on the natural unique keys."""
    import time
    from supabase import create_client

    outputs = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs")
    os.makedirs(outputs, exist_ok=True)
    client = create_client(url, key)

    def upsert(table, rows, conflict):
        n = 0
        for ch in _chunk(rows, 50):
            client.table(table).upsert(ch, on_conflict=conflict).execute()
            n += len(ch)
        return n

    def write_rejects(path, header, rows):
        import csv
        with open(path, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(header)
            w.writerows(rows)

    t0 = time.time()
    team_id = client.table("teams").select("id").eq("slug", "pier").single().execute().data["id"]
    today = dt.date.today().isoformat()  # matches the SQL path's date_added=CURRENT_DATE default
    timings = {}

    # --- Companies ---
    # NOT NULL columns with a DB default (date_added) must carry an explicit value:
    # PostgREST fills omitted keys with NULL, not the default, in multi-row upserts.
    t = time.time()
    crows = [_strip_none({**r, "team_id": team_id, "date_added": r.get("date_added") or today})
             for r in companies]
    comp_n = upsert("companies", crows, "company_id")
    timings["companies"] = time.time() - t

    # company_id (Cnnn) -> UUID map, pulled from what actually landed
    cmap = {}
    for ch in _chunk([r["company_id"] for r in companies], 200):
        for row in client.table("companies").select("id,company_id").in_("company_id", ch).execute().data:
            cmap[row["company_id"]] = row["id"]

    # --- Contacts (FK-validate company_ref) ---
    t = time.time()
    contact_rows, contact_rej = [], []
    for r in contacts:
        ref = r.get("company_ref")
        if not ref:
            contact_rej.append((r["contact_id"], "", "blank_company_ref")); continue
        if ref not in cmap:
            contact_rej.append((r["contact_id"], ref, "company_not_loaded")); continue
        # first_name/last_name are NOT NULL. Some workbook rows have a job_title +
        # company but no name (a research stub). Keep the row with '' rather than
        # dropping it (which would also orphan its outreach); flag for review.
        if not r.get("first_name") or not r.get("last_name"):
            warn("contacts.name (blank, defaulted to '')", r["contact_id"])
        contact_rows.append(_strip_none({**r, "team_id": team_id, "company_id": cmap[ref],
                                          "date_added": r.get("date_added") or today,
                                          "first_name": r.get("first_name") or "",
                                          "last_name": r.get("last_name") or ""}))
    con_n = upsert("contacts", contact_rows, "contact_id") if contact_rows else 0
    write_rejects(os.path.join(outputs, "rejected_contacts.csv"),
                  ["contact_id", "company_ref", "reason"], contact_rej)
    timings["contacts"] = time.time() - t

    # contact_id (Pnnn) -> (UUID, company UUID) map
    pmap = {}
    for ch in _chunk([r["contact_id"] for r in contact_rows], 200):
        for row in client.table("contacts").select("id,contact_id,company_id").in_("contact_id", ch).execute().data:
            pmap[row["contact_id"]] = (row["id"], row["company_id"])

    # --- Outreach (FK-validate contact_ref) ---
    t = time.time()
    out_rows, out_rej = [], []
    for r in outreach:
        ref = r.get("contact_ref")
        if not ref:
            out_rej.append((r["touch_id"], "", "blank_contact_ref")); continue
        if ref not in pmap:
            out_rej.append((r["touch_id"], ref, "contact_not_loaded")); continue
        cid, coid = pmap[ref]
        if not r.get("touch_date"):  # touch_date is NOT NULL with no DB default
            warn("outreach.touch_date (blank, defaulted to today)", r["touch_id"])
        out_rows.append(_strip_none({**r, "team_id": team_id, "contact_id": cid, "company_id": coid,
                                     "touch_date": r.get("touch_date") or today}))
    out_n = upsert("outreach_log", out_rows, "touch_id") if out_rows else 0
    write_rejects(os.path.join(outputs, "rejected_outreach.csv"),
                  ["touch_id", "contact_ref", "reason"], out_rej)
    timings["outreach"] = time.time() - t

    print("\n=== DIRECT LOAD COMPLETE ===")
    print(f"  companies    loaded {comp_n:>4}                              ({timings['companies']:.1f}s)")
    print(f"  contacts     loaded {con_n:>4}  rejected {len(contact_rej):>3}          ({timings['contacts']:.1f}s)")
    print(f"  outreach_log loaded {out_n:>4}  rejected {len(out_rej):>3}          ({timings['outreach']:.1f}s)")
    print(f"  total wall clock: {time.time() - t0:.1f}s")
    for tid, ref, why in contact_rej:
        print(f"    reject contact  {tid} ref={ref} {why}")
    for tid, ref, why in out_rej:
        print(f"    reject outreach {tid} ref={ref} {why}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    ap.add_argument("--phase", choices=["a", "b"], required=True)
    ap.add_argument("--emit-sql", help="write INSERT statements to this path")
    ap.add_argument("--dry-run", action="store_true", help="parse and report only")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        print(f"ERROR: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    companies = read_companies(wb["Companies"])

    if args.phase == "a":
        # Parse companies only, so the warning report reflects exactly what loads.
        by_id = {r["company_id"]: r for r in companies}
        missing = [cid for cid in PHASE_A_COMPANY_IDS if cid not in by_id]
        if missing:
            print(f"ERROR: Phase A ids not found in workbook: {missing}", file=sys.stderr)
            return 1
        companies = [by_id[cid] for cid in PHASE_A_COMPANY_IDS]
        contacts, outreach = [], []
        print(f"Phase A: {len(companies)} priority-weighted companies "
              f"(contacts/outreach deferred to phase B)")
    else:
        contacts = read_contacts(wb["Contacts"])
        outreach = read_outreach(wb["Outreach Log"])
        print(f"Phase B: {len(companies)} companies, {len(contacts)} contacts, {len(outreach)} outreach touches")

    if warnings:
        print("\nUnmapped / coerced values (value -> dropped or defaulted):")
        for k, n in warnings.most_common():
            print(f"  [{n:>4}x] {k}")
    else:
        print("\nNo mapping warnings.")

    if args.emit_sql:
        parts = [sql_companies(companies)] if companies else []
        if contacts:
            parts.append(sql_contacts(contacts))
        if outreach:
            parts.append(sql_outreach(outreach))
        with open(args.emit_sql, "w") as fh:
            fh.write("\n\n".join(parts))
        print(f"\nWrote SQL to {args.emit_sql}")

    if args.dry_run:
        print("--dry-run: nothing written to the database.")
        return 0

    if args.emit_sql:
        return 0

    # Default action: load directly via supabase-py (service role).
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY to load, "
              "or use --emit-sql / --dry-run.", file=sys.stderr)
        return 2
    return direct_load(companies, contacts, outreach, url, key)


if __name__ == "__main__":
    raise SystemExit(main())
